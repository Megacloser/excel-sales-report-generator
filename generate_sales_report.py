"""
Generate a polished, presentation-ready Excel sales report from a CSV file.

Usage:
    python generate_sales_report.py [input_csv] [output_xlsx]

Defaults: sample_sales_data.csv -> sales_report.xlsx
"""

import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Design tokens
# ---------------------------------------------------------------------------
PRIMARY = "1F3864"       # deep navy - brand color
PRIMARY_LIGHT = "DCE6F1"  # tinted navy for total row
BAND = "F2F5FA"           # zebra stripe
BORDER_COLOR = "D9D9D9"
TEXT = "1F1F1F"
SUBTLE = "AEB9CC"

MONEY_FORMAT = '$#,##0.00'
PERCENT_FORMAT = '0.0%'

TITLE_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=20)
SUBTITLE_FONT = Font(name="Calibri", color=SUBTLE, italic=True, size=10)
KPI_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(name="Calibri", color=TEXT, size=10.5)
TOTAL_FONT = Font(name="Calibri", color=PRIMARY, bold=True, size=11)

BANNER_FILL = PatternFill("solid", fgColor=PRIMARY)
HEADER_FILL = PatternFill("solid", fgColor=PRIMARY)
BAND_FILL = PatternFill("solid", fgColor=BAND)
TOTAL_FILL = PatternFill("solid", fgColor=PRIMARY_LIGHT)

THIN = Side(style="thin", color=BORDER_COLOR)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOTAL_BORDER = Border(left=THIN, right=THIN, bottom=THIN, top=Side(style="double", color=PRIMARY))

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def load_data(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    df.columns = [c.strip() for c in df.columns]
    required = {"Customer", "Product", "Amount", "Date"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%Y", errors="coerce")
    return df


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        df.groupby("Customer")
        .agg(**{"Orders": ("Product", "count"), "Total Spent": ("Amount", "sum")})
        .reset_index()
        .sort_values("Total Spent", ascending=False)
        .reset_index(drop=True)
    )
    grand_total = summary["Total Spent"].sum()
    summary.insert(0, "Rank", summary.index + 1)
    summary["% of Total"] = summary["Total Spent"] / grand_total if grand_total else 0
    return summary


def add_banner(ws, n_cols: int, title: str, subtitle: str, kpi_line: str):
    last_col = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = BANNER_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].fill = BANNER_FILL
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 18

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = kpi_line
    ws["A3"].font = KPI_FONT
    ws["A3"].fill = BANNER_FILL
    ws["A3"].alignment = CENTER
    ws.row_dimensions[3].height = 24

    ws.row_dimensions[4].height = 6


def autofit_columns(ws, n_cols: int, min_row: int, max_row: int, min_width: int = 10, padding: int = 3):
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col).value))
                for r in range(min_row, max_row + 1)
                if ws.cell(row=r, column=col).value is not None
            ),
            default=0,
        )
        ws.column_dimensions[letter].width = max(min_width, max_len + padding)


def write_summary_sheet(wb: Workbook, summary: pd.DataFrame):
    ws = wb.active
    ws.title = "Customer Report"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PRIMARY

    headers = list(summary.columns)
    n_cols = len(headers)
    last_col = get_column_letter(n_cols)

    total_orders = int(summary["Orders"].sum())
    total_revenue = float(summary["Total Spent"].sum())
    n_customers = len(summary)
    avg_order = total_revenue / total_orders if total_orders else 0

    add_banner(
        ws,
        n_cols,
        title="SALES REPORT",
        subtitle=f"Customer Summary  •  Generated {date.today():%B %d, %Y}",
        kpi_line=(
            f"Total Revenue: ${total_revenue:,.2f}   |   Orders: {total_orders}   |   "
            f"Customers: {n_customers}   |   Avg Order Value: ${avg_order:,.2f}"
        ),
    )

    header_row = 5
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    ws.row_dimensions[header_row].height = 20

    first_data_row = header_row + 1
    for i, (_, row) in enumerate(summary.iterrows()):
        r = first_data_row + i
        values = [
            int(row["Rank"]),
            row["Customer"],
            int(row["Orders"]),
            float(row["Total Spent"]),
            float(row["% of Total"]),
        ]
        fill = BAND_FILL if i % 2 == 1 else None
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            if fill:
                cell.fill = fill
            if c in (1, 3):
                cell.alignment = CENTER
            elif c == 4:
                cell.number_format = MONEY_FORMAT
                cell.alignment = RIGHT
            elif c == 5:
                cell.number_format = PERCENT_FORMAT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT

    last_data_row = first_data_row + len(summary) - 1
    total_row = last_data_row + 2  # blank spacer row keeps TOTAL out of the filter range

    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    total_values = {1: "TOTAL", 3: total_orders, 4: total_revenue, 5: 1.0}
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=total_row, column=c, value=total_values.get(c))
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = TOTAL_BORDER
        if c == 1:
            cell.alignment = LEFT
        elif c == 3:
            cell.alignment = CENTER
        elif c == 4:
            cell.number_format = MONEY_FORMAT
            cell.alignment = RIGHT
        elif c == 5:
            cell.number_format = PERCENT_FORMAT
            cell.alignment = RIGHT
    ws.row_dimensions[total_row].height = 20

    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    autofit_columns(ws, n_cols, header_row, total_row)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Total Spent by Customer"
    chart.y_axis.title = "Amount, $"
    chart.x_axis.title = "Customer"
    chart.style = 10
    chart.legend = None
    chart.gapWidth = 60

    data = Reference(ws, min_col=4, min_row=header_row, max_row=last_data_row)
    cats = Reference(ws, min_col=2, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = PRIMARY
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.width = 20
    chart.height = 10
    ws.add_chart(chart, f"A{total_row + 3}")

    return ws


def write_data_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Sales Data")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C9A227"

    headers = ["Customer", "Product", "Amount", "Date"]
    n_cols = len(headers)
    last_col = get_column_letter(n_cols)

    add_banner(
        ws,
        n_cols,
        title="SALES DATA",
        subtitle=f"Full Transaction Log  •  Generated {date.today():%B %d, %Y}",
        kpi_line=f"{len(df)} transactions",
    )

    header_row = 5
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    ws.row_dimensions[header_row].height = 20

    first_data_row = header_row + 1
    for i, (_, row) in enumerate(df.iterrows()):
        r = first_data_row + i
        values = [
            row["Customer"],
            row["Product"],
            float(row["Amount"]),
            row["Date"].strftime("%m/%d/%Y") if pd.notna(row["Date"]) else "",
        ]
        fill = BAND_FILL if i % 2 == 1 else None
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            if fill:
                cell.fill = fill
            if c == 3:
                cell.number_format = MONEY_FORMAT
                cell.alignment = RIGHT
            elif c == 4:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    last_data_row = first_data_row + len(df) - 1
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    autofit_columns(ws, n_cols, header_row, last_data_row)


def build_report(input_csv: str, output_xlsx: str):
    csv_path = Path(input_csv)
    if not csv_path.exists():
        raise FileNotFoundError(f"File not found: {csv_path}")

    df = load_data(csv_path)
    summary = build_summary(df)

    wb = Workbook()
    write_summary_sheet(wb, summary)
    write_data_sheet(wb, df)

    out_path = Path(output_xlsx)
    wb.save(out_path)
    print(f"Report saved: {out_path.resolve()}")
    print(f"Customers: {len(summary)}, orders: {len(df)}, total revenue: ${summary['Total Spent'].sum():,.2f}")


if __name__ == "__main__":
    input_csv = sys.argv[1] if len(sys.argv) > 1 else "sample_sales_data.csv"
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else "sales_report.xlsx"
    build_report(input_csv, output_xlsx)
